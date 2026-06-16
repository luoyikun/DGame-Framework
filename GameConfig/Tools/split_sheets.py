# -*- coding: utf-8 -*-
"""
split_sheets.py —— Luban 导表前置 / 后置预处理脚本

背景：
    Luban 4.4.1+ 原生支持「自动导入表」（DefaultTableImporter）：
    扫描配置目录下文件名以 `#` 开头的 Excel（如 `#ItemCfg-道具表.xlsx`），
    自动注册成表（`#` 后→value_type，加 Tb 前缀→full_name，`-` 后→comment）。
    该能力是【文件级】的，只看「文件名」，不看 sheet 名。

    但 Luban 的模型是「一个 # 文件 = 一张表」：同一个文件里的多个 sheet 会被
    【合并】成同一张表的数据（实测会触发主键冲突），无法让每个 sheet 各自成表；
    且本项目使用的 Luban 构建不支持 `文件.xlsx@sheet` 语法。

    因此本脚本只负责补齐【sheet 级】能力：
    把任意 Excel 文件中 sheet 名匹配 `#xxx-xxx` 的 sheet，
    各自拆成一个独立的单 sheet 临时文件 `#<value_type>-<comment>.xlsx`，
    放到 Datas 根目录，交给 Luban 原生 auto-import 去成表。

    文件级（文件名 `#xxx-xxx.xlsx`）由 Luban 原生兜住，本脚本不处理。

命名规则（与 Luban 原生一致）：
    sheet 名 `#<value_type>-<comment>`
        `#` 后、`-` 前   -> value_type（Luban 自动加 Tb 前缀成 full_name）
        `-` 后           -> comment（可选，不写则无注释）
    例：`#SkillCfg-技能表` -> value_type=SkillCfg, comment=技能表 -> 生成 TbSkillCfg

    注意：拆分产物文件名必须是标准 `#value_type-comment.xlsx`，不能带额外的 `.`，
    因为 Luban 会把文件名里的 `.` 当成 namespace 分隔符（实测 `#Pet.split.xlsx`
    会被解析成 namespace=Pet、类名=split）。因此本脚本【不在文件名里加标记】，
    改用清单文件 `_split_manifest.txt` 精确记录本轮生成的产物，供清理使用。

去重 / 不覆盖：
    若某 value_type 已存在（Datas 根目录已有用户手写的同名 `#value_type*.xlsx`，
    或本轮已生成过同名），则跳过，不重复生成、不覆盖任何已有内容。

用法：
    生成： python split_sheets.py --datas <Datas目录>
    清理： python split_sheets.py --datas <Datas目录> --clean

    do_split 会在生成前先按清单清掉上一轮残留，保证可重复执行（幂等）。
"""

import argparse
import os
import re
import sys

try:
    import openpyxl
except ImportError:
    sys.stderr.write("[split_sheets] 缺少依赖 openpyxl，请先执行: pip install openpyxl\n")
    sys.exit(1)

# sheet 名匹配规则：# 开头，# 后为 value_type，可选 - 后为 comment
SHEET_PATTERN = re.compile(r"^#([^\-]+)(?:-(.*))?$")

# Luban schema 表，不参与拆分
SCHEMA_FILES = {"__tables__", "__beans__", "__enums__"}

# 清单文件名（记录本脚本生成的拆分产物，供清理；非 .xlsx，不会被 Luban 当数据表）
MANIFEST_NAME = "_split_manifest.txt"


def manifest_path(datas_dir):
    return os.path.join(datas_dir, MANIFEST_NAME)


def read_manifest(datas_dir):
    p = manifest_path(datas_dir)
    if not os.path.isfile(p):
        return []
    with open(p, "r", encoding="utf-8") as f:
        return [line.strip() for line in f if line.strip()]


def write_manifest(datas_dir, names):
    with open(manifest_path(datas_dir), "w", encoding="utf-8") as f:
        for n in names:
            f.write(n + "\n")


def remove_generated(datas_dir):
    """按清单删除上一轮生成的拆分产物及清单本身，返回删除个数。"""
    removed = 0
    for name in read_manifest(datas_dir):
        fp = os.path.join(datas_dir, name)
        if os.path.isfile(fp):
            try:
                os.remove(fp)
                removed += 1
            except OSError as e:
                sys.stderr.write(f"[split_sheets] 删除失败 {name}: {e}\n")
    mp = manifest_path(datas_dir)
    if os.path.isfile(mp):
        try:
            os.remove(mp)
        except OSError:
            pass
    return removed


def iter_source_xlsx(datas_dir):
    """收集 Datas 下所有业务 Excel（跳过临时文件、schema 表）。

    先收集成 list 再返回，避免边遍历边生成产物时把刚生成的文件又扫进来。
    """
    result = []
    for root, _dirs, files in os.walk(datas_dir):
        for fn in files:
            if not fn.lower().endswith(".xlsx"):
                continue
            if fn.startswith("~$"):          # Excel 打开时的锁定临时文件
                continue
            stem = os.path.splitext(fn)[0]
            if stem in SCHEMA_FILES:
                continue
            result.append(os.path.join(root, fn))
    return result


def collect_existing_value_types(datas_dir):
    """收集 Datas 根目录下用户手写 `#` 文件对应的 value_type，用于去重。

    Luban 原生 auto-import 已会处理文件名为 `#xxx` 的文件，
    若 value_type 已有同名 `#` 文件，则 sheet 拆分应跳过，避免重复成表 / 主键冲突。
    调用前需先 remove_generated()，确保此时 Datas 里的 `#` 文件都是用户手写的。
    """
    existing = set()
    for fn in os.listdir(datas_dir):
        if not fn.lower().endswith(".xlsx") or fn.startswith("~$"):
            continue
        if fn.startswith("#"):
            stem = os.path.splitext(fn)[0]
            m = SHEET_PATTERN.match(stem)
            if m:
                existing.add(m.group(1).strip())
    return existing


def parse_sheet_name(sheet_name):
    """解析 sheet 名，返回 (value_type, comment) 或 None（不匹配）。"""
    m = SHEET_PATTERN.match(sheet_name.strip())
    if not m:
        return None
    value_type = m.group(1).strip()
    comment = (m.group(2) or "").strip()
    if not value_type:
        return None
    return value_type, comment


def safe_filename_part(text):
    """清理 comment 中不适合做文件名的字符（含 . 以免被 Luban 当 namespace 分隔符）。"""
    return re.sub(r'[\\/:*?"<>|.]', "_", text)


def split_one_sheet(src_path, sheet_name, out_path):
    """把 src 文件里指定 sheet 复制成一个独立单 sheet 工作簿，存到 out_path。"""
    wb_src = openpyxl.load_workbook(src_path, data_only=False)
    ws_src = wb_src[sheet_name]

    wb_dst = openpyxl.Workbook()
    ws_dst = wb_dst.active
    ws_dst.title = "Sheet1"  # 文件级 auto-import 只看文件名，sheet 名无所谓

    for row in ws_src.iter_rows(values_only=True):
        ws_dst.append(list(row))

    wb_dst.save(out_path)
    wb_src.close()


def do_split(datas_dir):
    # 先清掉上一轮残留，保证幂等；之后 Datas 里的 # 文件都是用户手写的
    remove_generated(datas_dir)

    existing_value_types = collect_existing_value_types(datas_dir)
    generated_value_types = set()   # 本轮已生成的 value_type，避免不同文件同名 sheet 互相覆盖
    manifest = []                   # 本轮生成的文件名，写入清单供清理
    created, skipped = [], []

    for src_path in iter_source_xlsx(datas_dir):
        src_name = os.path.basename(src_path)
        try:
            wb = openpyxl.load_workbook(src_path, read_only=True)
            sheet_names = wb.sheetnames
            wb.close()
        except Exception as e:
            sys.stderr.write(f"[split_sheets] 跳过无法读取的文件 {src_name}: {e}\n")
            continue

        for sn in sheet_names:
            parsed = parse_sheet_name(sn)
            if parsed is None:
                continue
            value_type, comment = parsed

            # 去重：已有同名 # 文件 / 已被本轮生成 -> 跳过，不覆盖
            if value_type in existing_value_types or value_type in generated_value_types:
                skipped.append((src_name, sn, value_type, "已存在同名表"))
                continue

            if comment:
                out_name = f"#{value_type}-{safe_filename_part(comment)}.xlsx"
            else:
                out_name = f"#{value_type}.xlsx"
            out_path = os.path.join(datas_dir, out_name)

            try:
                split_one_sheet(src_path, sn, out_path)
            except Exception as e:
                sys.stderr.write(f"[split_sheets] 拆分失败 {src_name}[{sn}]: {e}\n")
                continue

            generated_value_types.add(value_type)
            manifest.append(out_name)
            created.append((src_name, sn, out_name))

    if manifest:
        write_manifest(datas_dir, manifest)

    print(f"[split_sheets] sheet 拆分完成：新增 {len(created)} 个，跳过 {len(skipped)} 个")
    for src_name, sn, out_name in created:
        print(f"  + {src_name} [{sn}] -> {out_name}")
    for src_name, sn, value_type, reason in skipped:
        print(f"  - {src_name} [{sn}] (value_type={value_type}) 跳过：{reason}")


def do_clean(datas_dir):
    removed = remove_generated(datas_dir)
    print(f"[split_sheets] 清理完成：删除 {removed} 个临时拆分文件")


def main():
    parser = argparse.ArgumentParser(description="Luban 导表 sheet 级自动注册预处理")
    parser.add_argument("--datas", required=True, help="Datas 目录路径")
    parser.add_argument("--clean", action="store_true", help="清理上次生成的拆分临时文件")
    args = parser.parse_args()

    datas_dir = os.path.abspath(args.datas)
    if not os.path.isdir(datas_dir):
        sys.stderr.write(f"[split_sheets] Datas 目录不存在: {datas_dir}\n")
        sys.exit(1)

    if args.clean:
        do_clean(datas_dir)
    else:
        do_split(datas_dir)


if __name__ == "__main__":
    main()
